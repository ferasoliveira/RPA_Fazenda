import os
import shutil
from pathlib import Path
import openpyxl
from openpyxl import Workbook, load_workbook

# Base directory for processed data
BASE_DIR = "Dados Processados"

# Mapping from abbreviated farm names to full folder names
FARM_MAPPING = {
    "VAL": "Valongo",
    "BARRA": "Barra",
    "TRES": "Três Divisas"
}

def ensure_full_path(farm_code, batch_name, date_str):
    """
    Creates the full path: Dados Processados / [Fazenda] / [Lote] / [Data]
    """
    farm_name = FARM_MAPPING.get(farm_code, farm_code)
    path = Path(BASE_DIR) / farm_name / batch_name / date_str
    path.mkdir(parents=True, exist_ok=True)
    return path

def get_excel_path(full_path, date_str):
    """
    Returns the path for the Excel file inside the target folder.
    """
    return full_path / f"{date_str}.xlsx"

def check_duplicate(excel_path, brinco_number):
    """
    Checks if the brinco_number already exists in the Excel file using openpyxl.
    """
    if not excel_path.exists():
        return False
    
    try:
        wb = load_workbook(excel_path)
        sheet = wb.active
        
        # Look for "BRINCOS" in the first row
        brincos_col = 1
        for cell in sheet[1]:
            if cell.value == "BRINCOS":
                brincos_col = cell.column
                break
        
        for row in range(2, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=brincos_col).value
            if str(cell_value) == str(brinco_number):
                wb.close()
                return True
        wb.close()
    except Exception as e:
        print(f"Error checking duplicates: {e}")
    
    return False

def add_to_excel(excel_path, brinco_number):
    """
    Adds the brinco_number to the Excel file using openpyxl. Creates it if it doesn't exist.
    """
    if excel_path.exists():
        wb = load_workbook(excel_path)
        sheet = wb.active
    else:
        wb = Workbook()
        sheet = wb.active
        sheet.cell(row=1, column=1, value="BRINCOS")
    
    next_row = sheet.max_row + 1
    sheet.cell(row=next_row, column=1, value=brinco_number)
    
    wb.save(excel_path)
    wb.close()

def remove_last_from_excel(excel_path, brinco_number):
    """
    Removes the last entry for a specific brinco_number from the Excel file.
    """
    if not excel_path.exists():
        return False
    
    try:
        wb = load_workbook(excel_path)
        sheet = wb.active
        
        # Look for "BRINCOS" in the first row
        brincos_col = 1
        for cell in sheet[1]:
            if cell.value == "BRINCOS":
                brincos_col = cell.column
                break
        
        # Iterate backwards to find the last occurrence
        for row in range(sheet.max_row, 1, -1):
            if str(sheet.cell(row=row, column=brincos_col).value) == str(brinco_number):
                sheet.delete_rows(row)
                wb.save(excel_path)
                wb.close()
                return True
        wb.close()
    except Exception as e:
        print(f"Error removing from excel: {e}")
    
    return False

def process_photo(photo_path, target_folder, brinco_number, excel_path):
    """
    Moves, renames the photo, and adds the data to Excel.
    """
    # 1. Check for duplicates in Excel
    if check_duplicate(excel_path, brinco_number):
        return False, f"O número {brinco_number} já foi adicionado."

    # 2. Add to Excel
    try:
        add_to_excel(excel_path, brinco_number)
    except Exception as e:
        return False, f"Erro ao atualizar Excel: {e}"

    # 3. Rename and Move photo
    photo_ext = Path(photo_path).suffix
    new_photo_name = f"{brinco_number}{photo_ext}"
    destination = target_folder / new_photo_name
    
    # Handle if file already exists in destination
    if destination.exists():
        return False, f"Já existe uma foto com o nome {new_photo_name} na pasta de destino."

    try:
        shutil.move(photo_path, destination)
    except Exception as e:
        # Rollback Excel if move fails
        remove_last_from_excel(excel_path, brinco_number)
        return False, f"Erro ao mover foto: {e}"

    return True, "Sucesso"

def undo_process(photo_info):
    """
    Reverts the process for one photo.
    """
    try:
        remove_last_from_excel(Path(photo_info['excel_path']), photo_info['brinco_number'])
        dest = Path(photo_info['dest_path'])
        orig = Path(photo_info['original_path'])
        if dest.exists():
            shutil.move(dest, orig)
            return True, "Desfeito com sucesso"
        else:
            return False, "Arquivo original não encontrado para restaurar"
    except Exception as e:
        return False, f"Erro ao desfazer: {e}"

def list_photos_to_process(source_dir="processar"):
    """
    Lists all image files in the processar directory.
    """
    source_path = Path(source_dir)
    if not source_path.exists():
        source_path.mkdir(parents=True, exist_ok=True)
        return []
    
    extensions = ('.png', '.jpg', '.jpeg', '.bmp', '.gif')
    return [str(f) for f in sorted(source_path.iterdir()) if f.suffix.lower() in extensions]

def levenshtein_distance(s1, s2):
    """
    Calculates the Levenshtein distance between two strings.
    """
    if len(s1) < len(s2):
        return levenshtein_distance(s2, s1)
    if len(s2) == 0:
        return len(s1)
    previous_row = range(len(s2) + 1)
    for i, c1 in enumerate(s1):
        current_row = [i + 1]
        for j, c2 in enumerate(s2):
            insertions = previous_row[j + 1] + 1
            deletions = current_row[j] + 1
            substitutions = previous_row[j] + (c1 != c2)
            current_row.append(min(insertions, deletions, substitutions))
        previous_row = current_row
    return previous_row[-1]

def search_brinco(query):
    """
    Searches for a brinco in all processed folders.
    Returns (exact_matches, near_matches)
    """
    exact_matches = []
    near_matches = []
    query = str(query).strip()
    
    base_path = Path(BASE_DIR)
    if not base_path.exists():
        return [], []

    # Traverse Dados Processados / [Farm] / [Batch] / [Date]
    for farm_dir in base_path.iterdir():
        if not farm_dir.is_dir(): continue
        for batch_dir in farm_dir.iterdir():
            if not batch_dir.is_dir(): continue
            for date_dir in batch_dir.iterdir():
                if not date_dir.is_dir(): continue
                excel_files = list(date_dir.glob("*.xlsx"))
                for excel_file in excel_files:
                    try:
                        wb = load_workbook(excel_file, read_only=True)
                        sheet = wb.active
                        for row in range(2, sheet.max_row + 1):
                            val = str(sheet.cell(row=row, column=1).value).strip()
                            if val == query:
                                exact_matches.append({
                                    'farm': farm_dir.name,
                                    'batch': batch_dir.name,
                                    'date': date_dir.name,
                                    'brinco': val
                                })
                            elif levenshtein_distance(val, query) == 1:
                                near_matches.append({
                                    'farm': farm_dir.name,
                                    'batch': batch_dir.name,
                                    'date': date_dir.name,
                                    'brinco': val
                                })
                        wb.close()
                    except Exception:
                        continue
    
    # If exact matches are found, return only them (don't show near matches)
    if exact_matches:
        return exact_matches, []
    
    return [], near_matches
